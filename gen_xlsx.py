import openpyxl
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
from openpyxl.worksheet.datavalidation import DataValidation
import datetime
import os
import sys
import json

# ─── Formatting helpers ──────────────────────────────────────────────────────
def thin_border():
    return Border(left=Side(style='thin'), right=Side(style='thin'),
                  top=Side(style='thin'), bottom=Side(style='thin'))

def fill(color_hex):
    return PatternFill(start_color=color_hex, end_color=color_hex, fill_type='solid')

CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)
LEFT = Alignment(horizontal='left', vertical='center', wrap_text=True)

# ─── Header builder ──────────────────────────────────────────────────────────
def build_header(ws, task_id: str, data_len: int = 0):
    B = thin_border()
    yellow = fill('FFFF99')
    today = datetime.date.today().strftime('%m/%d/%y')

    # branding col A
    ws.merge_cells('A1:A4')
    ws['A1'].fill = fill('D9EAD3'); ws['A1'].border = B

    # System / Subsystem
    ws.merge_cells('B1:D2'); ws['B1'].value = 'システム'; ws['B1'].alignment = CENTER
    ws.merge_cells('E1:J2'); ws['E1'].value = 'サブシステム'; ws['E1'].alignment = CENTER
    for c in range(2, 11):
        for r in range(1, 3): ws.cell(r, c).border = B

    # ID blocks
    ws.merge_cells('B3:B4'); ws['B3'].value = 'ID'; ws['B3'].fill = yellow; ws['B3'].alignment = CENTER; ws['B3'].border = B
    ws.merge_cells('C3:D4'); ws['C3'].value = task_id; ws['C3'].alignment = CENTER; ws['C3'].border = B
    ws.merge_cells('E3:E4'); ws['E3'].value = 'ID'; ws['E3'].fill = yellow; ws['E3'].alignment = CENTER; ws['E3'].border = B
    ws.merge_cells('F3:H4'); ws['F3'].value = '工程写真アプリ / かん助'; ws['F3'].alignment = CENTER; ws['F3'].border = B
    ws.merge_cells('I3:I4'); ws['I3'].value = 'ID'; ws['I3'].fill = yellow; ws['I3'].alignment = CENTER; ws['I3'].border = B
    ws.merge_cells('J3:J4'); ws['J3'].alignment = CENTER; ws['J3'].border = B
    for r in range(3, 5):
        for c in range(2, 11): ws.cell(r, c).border = B

    # Instructions K-P
    ws.merge_cells('K1:P4')
    ws['K1'].value = '・A列でLvを選択する事で色付け可能\n・I列【不要】・・・テスト件数に含めるかどうか\n\u3000\u3000●：含める\n\u3000\u3000不：含めない'
    ws['K1'].font = Font(color='FF0000'); ws['K1'].alignment = Alignment(wrap_text=True, vertical='top'); ws['K1'].border = B

    # Metadata Q-U
    ws['Q1'].value = '処理'; ws['Q1'].fill = yellow; ws['Q1'].alignment = CENTER; ws['Q1'].border = B
    ws.merge_cells('R1:S1'); ws['R1'].value = '日付'; ws['R1'].fill = yellow; ws['R1'].border = B
    ws.merge_cells('T1:U1'); ws['T1'].value = '担当'; ws['T1'].fill = yellow; ws['T1'].border = B
    ws['Q2'].value = '作成'; ws['Q2'].fill = yellow; ws['Q2'].border = B
    ws.merge_cells('R2:S2'); ws['R2'].value = today; ws['R2'].border = B
    ws.merge_cells('T2:U2'); ws['T2'].value = 'RS.HCM'; ws['T2'].border = B
    ws['Q3'].value = '更新'; ws['Q3'].fill = yellow; ws['Q3'].border = B
    ws.merge_cells('R3:S3'); ws['R3'].border = B
    ws.merge_cells('T3:U3'); ws['T3'].border = B
    for c in range(17, 22): ws.cell(4, c).border = B

    # Statistics X-AH Row 2
    stats = [('X', 'IT指摘', 'Y', '0'), ('Z', 'テスト項目', 'AA', str(data_len)),
             ('AB', 'テスト消化', 'AC', '0%'), ('AD', 'ステップ数', 'AE', 'Step'), ('AF', '障害件数', 'AG', '')]
    for lc, lt, vc, vt in stats:
        ws[f'{lc}2'].value = lt; ws[f'{lc}2'].fill = yellow; ws[f'{lc}2'].border = B; ws[f'{lc}2'].alignment = CENTER
        ws[f'{vc}2'].value = vt; ws[f'{vc}2'].border = B; ws[f'{vc}2'].alignment = CENTER
    ws.merge_cells('AG2:AH2'); ws['AG2'].border = B

def write_table_headers(ws):
    B = thin_border(); h_fill = fill('CCFFCC')
    headers_56 = [('A', 'Lv'), ('B', '項目\nNo'), ('C', '評価項目'), ('D', 'テストケース\nNo'), 
                  ('E', '対象アプリ'), ('F', 'テストケース内容'), ('G', '判定結果（予想結果）'), ('H', '検証環境'), ('I', '不要')]
    for col, txt in headers_56:
        ws.merge_cells(f'{col}5:{col}6')
        ws[f'{col}5'].value = txt; ws[f'{col}5'].fill = h_fill; ws[f'{col}5'].alignment = CENTER; ws[f'{col}5'].border = B
    
    # Tests merge
    ws.merge_cells('J5:L5'); ws['J5'].value = 'テスト実施 (1回目)'; ws['J5'].fill = h_fill; ws['J5'].border = B; ws['J5'].alignment = CENTER
    ws.merge_cells('M5:O5'); ws['M5'].value = 'テスト実施 (2回目)'; ws['M5'].fill = h_fill; ws['M5'].border = B; ws['M5'].alignment = CENTER
    for i, l in enumerate(['合否','実施者','実施日']*2, 10):
        c = ws.cell(6, i); c.value = l; c.fill = h_fill; c.border = B; c.alignment = CENTER
    
    # Extras
    for i, (col, txt) in enumerate([('P','備考'),('Q','観点'),('R','デグレ除く'),('S','デグレ'),('T','テスト総数'),('U','テスト消化率')], 16):
        ws.merge_cells(f'{col}5:{col}6')
        c = ws[f'{col}5']; c.value = txt; c.fill = h_fill; c.border = B; c.alignment = CENTER

def write_data_rows(ws, data):
    B = thin_border()
    for r, item in enumerate(data, 7):
        mapping = [('A','lv'),('B','item_no'),('C','category'),('D','tc_no'),('E','app'),('F','procedure'),('G','expected'),('H','env'),('I','fuyou')]
        for col, key in mapping:
            c = ws[f'{col}{r}']; c.value = item.get(key, ''); c.border = B; c.alignment = LEFT if col in 'FG' else CENTER
            if item.get('highlight') and col == 'D': c.font = Font(color='0000FF', bold=True)
        for i in range(10, 22): ws.cell(r, i).border = B
    return 7 + len(data) - 1

def generate_xlsx(task_id: str, output_dir: str = '.'):
    # 1. Load Data from task_id/data.json
    data_path = os.path.join(task_id, "data.json")
    if not os.path.exists(data_path):
        print(f"Error: Missing data.json in {task_id}"); return
    with open(data_path, 'r', encoding='utf-8') as f: data = json.load(f)

    # 2. Load History
    hist_path = os.path.join(task_id, "history.json")
    hist = []
    if os.path.exists(hist_path):
        with open(hist_path, 'r', encoding='utf-8') as f: hist = json.load(f)
    hist.sort(key=lambda x: x.get('datetime', ''))

    wb = Workbook()
    
    # Sheet 履歴
    ws_h = wb.active; ws_h.title = '履歴'; h_fill = fill('DDDDDD'); B = thin_border()
    for i, l in enumerate(['更新日時', '更新内容', '対象範囲'], 1):
        c = ws_h.cell(1, i, l); c.fill = h_fill; c.border = B; c.alignment = CENTER
    for r, h in enumerate(hist, 2):
        ws_h.cell(r,1,h.get('datetime','')); ws_h.cell(r,2,h.get('note','')); ws_h.cell(r,3,h.get('range',''))
        for i in range(1,4): ws_h.cell(r,i).border = B
    ws_h.column_dimensions['A'].width = 20; ws_h.column_dimensions['B'].width = 60

    # Sheet チェックリスト
    ws = wb.create_sheet('チェックリスト')
    build_header(ws, task_id, len(data))
    write_table_headers(ws)
    last_row = write_data_rows(ws, data)
    
    # Widths
    widths = {'A':5,'B':6,'C':20,'D':10,'E':15,'F':45,'G':45,'H':15,'I':5,'P':30}
    for k, v in widths.items(): ws.column_dimensions[k].width = v
    
    wb.move_sheet('チェックリスト', offset=1)
    os.makedirs(output_dir, exist_ok=True)
    out = os.path.join(output_dir, f'KANSUKE_Checklist_{task_id}.xlsx')
    wb.save(out); print(f'>>> Generated: {out}')

if __name__ == '__main__':
    if len(sys.argv) < 2: print("Usage: python gen_xlsx.py <task_id>"); sys.exit(1)
    generate_xlsx(sys.argv[1], sys.argv[2] if len(sys.argv)>2 else '.')
